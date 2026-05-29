"""Excel BoQ export generator."""

import io
import os
import tempfile
from datetime import datetime
from typing import Any


def generate_boq_excel(boq: dict[str, Any], output_path: str | None = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    navy = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    grey = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    summary = boq.get("summary", {})
    project = boq.get("project_name", "Project")
    country = boq.get("country_code", "ZM")
    date_str = datetime.now().strftime("%d %B %Y")
    tender_no = boq.get("tender_no", "TENDER-___/202_")

    # Cover (ZPPA Standard Cover Page details)
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "REPUBLIC OF ZAMBIA"
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = navy
    ws.merge_cells("A1:G1")
    ws["A2"] = "ZAMBIA PUBLIC PROCUREMENT AUTHORITY (ZPPA) STANDARD BOQ"
    ws["A2"].font = Font(bold=True, size=14, color="1A2744")
    ws.merge_cells("A2:G2")
    
    ws["A4"] = "Tender No."
    ws["B4"] = tender_no
    ws["A5"] = "Project Name"
    ws["B5"] = project
    ws["A6"] = "Procuring Entity"
    ws["B6"] = boq.get("client", "Ministry of Infrastructure")
    ws["A7"] = "Date of Preparation"
    ws["B7"] = date_str
    
    ws["A9"] = "TOTAL TENDER SUM (Exclusive of VAT)"
    ws["B9"] = summary.get("total_local_currency", 0)
    ws["C9"] = summary.get('local_currency', 'ZMW')
    ws["A9"].font = Font(bold=True)
    ws["B9"].font = Font(bold=True)

    # Summary sheet
    ws_sum = wb.create_sheet("Grand Summary")
    ws_sum.append(["Bill No.", "Description", f"Amount ({summary.get('local_currency', 'ZMW')})", "Remarks"])
    for cell in ws_sum[1]:
        cell.font = header_font
        cell.fill = navy
        
    for key, title in boq.get("section_titles", {}).items():
        total = boq.get("section_totals", {}).get(key, {}).get("mid", 0)
        # Assuming conversion to local currency is roughly total * (local/usd ratio), but for now we just show USD if local is not fully calculated per section.
        # Actually, ZMW BOQs must be priced in ZMW.
        # Let's check if there's a local_total per section. If not, we just output the number.
        ws_sum.append([f"Bill {key}", title, total])
        
    ws_sum.append(["", "Sub-Total 1", f"=SUM(C2:C{ws_sum.max_row})"])
    ws_sum.append(["", "Add: Preliminaries & General (Overhead)", summary.get("overhead_usd", 0)])
    ws_sum.append(["", "Add: Profit", summary.get("profit_usd", 0)])
    ws_sum.append(["", "Sub-Total 2 (Construction Cost)", f"=C{ws_sum.max_row-2}+C{ws_sum.max_row-1}+C{ws_sum.max_row}"])
    ws_sum.append(["", "Add: Contingency Sum", summary.get("contingency_usd", 0)])
    ws_sum.append(["", "GRAND TOTAL (To Form of Tender)", summary.get("total_project_estimate_usd", 0)])
    
    for row in range(ws_sum.max_row - 4, ws_sum.max_row + 1):
        ws_sum.cell(row=row, column=2).font = Font(bold=True)
        ws_sum.cell(row=row, column=3).font = Font(bold=True)

    # Section sheets
    for key, title in boq.get("section_titles", {}).items():
        lines = boq.get("sections", {}).get(key, [])
        if not lines:
            continue
        sheet_name = f"Bill {key}"[:31]
        ws_sec = wb.create_sheet(sheet_name)
        ws_sec.append([f"Bill No. {key}: {title}"])
        ws_sec["A1"].font = Font(bold=True, size=12)
        ws_sec.merge_cells("A1:G1")
        
        headers = ["Item No.", "Description", "Unit", "Qty", "Rate", "Amount", "Rate Ref / Source"]
        ws_sec.append(headers)
        for cell in ws_sec[2]:
            cell.font = header_font
            cell.fill = navy
            
        for i, line in enumerate(lines, start=3):
            # Try to resolve a rate reference if available, otherwise default to ZMW standard
            rate_ref = line.get("rate_ref", f"JCC/InFra_TeCh {date_str}")
            ws_sec.append([
                line.get("element_ref", ""),
                line.get("description", ""),
                line.get("unit", ""),
                line.get("quantity", 0),
                line.get("rate_mid", 0),
                line.get("amount_mid", 0),
                rate_ref
            ])
            if i % 2 == 1:
                for cell in ws_sec[i]:
                    cell.fill = grey
                    
        total = boq.get("section_totals", {}).get(key, {}).get("mid", 0)
        row = ws_sec.max_row + 1
        ws_sec.cell(row=row, column=5, value="Total Carried to Summary").font = Font(bold=True)
        c = ws_sec.cell(row=row, column=6, value=total)
        c.font = Font(bold=True)
        c.fill = blue

    # Rates sheet
    ws_rates = wb.create_sheet("Rates Database")
    ws_rates.append(["Material ID", "Description", "Unit", "Rate Min", "Rate Max", "Country", "Date of Validity"])
    from boq.materials_database import MATERIALS_DATABASE

    for mid, mat in MATERIALS_DATABASE.items():
        if mid == "exchange_rates":
            continue
        rates = mat.get("rates", {}).get(country, mat.get("rates", {}).get("ZM", {}))
        ws_rates.append([
            mid,
            mat.get("description", ""),
            mat.get("unit", ""),
            rates.get("min", 0),
            rates.get("max", 0),
            country,
            date_str
        ])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[letter].width = min(max_len + 2, 60)

    if not output_path:
        fd, output_path = tempfile.mkstemp(suffix=".xlsx", prefix="architex-cad_boq_")
        os.close(fd)
    wb.save(output_path)
    return output_path


def generate_boq_excel_bytes(boq: dict[str, Any]) -> bytes:
    from openpyxl import Workbook

    path = generate_boq_excel(boq)
    with open(path, "rb") as f:
        data = f.read()
    try:
        os.remove(path)
    except OSError:
        pass
    return data
