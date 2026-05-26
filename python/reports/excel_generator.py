"""Excel calculation report generator."""

from typing import Any


def generate_excel_report(calculation_result: dict[str, Any], output_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation Report"

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")

    ws["A1"] = "ARCHITEX-CAD Calculation Report"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:D1")

    ws["A3"] = "Status"
    ws["B3"] = calculation_result.get("status", "").upper()
    ws["A4"] = "Timestamp"
    ws["B4"] = calculation_result.get("timestamp", "")

    row = 6
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True)
    row += 1
    for key, value in calculation_result.get("summary", {}).items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Calculation Steps").font = Font(bold=True)
    row += 1
    headers = ["Step", "Title", "Formula", "Substitution", "Result", "Reference"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
    row += 1

    for step in calculation_result.get("steps", []):
        ws.cell(row=row, column=1, value=step.get("step_number"))
        ws.cell(row=row, column=2, value=step.get("title"))
        ws.cell(row=row, column=3, value=step.get("formula"))
        ws.cell(row=row, column=4, value=step.get("substitution"))
        ws.cell(row=row, column=5, value=step.get("result"))
        ws.cell(row=row, column=6, value=step.get("reference", ""))
        row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output_path)
    return output_path
